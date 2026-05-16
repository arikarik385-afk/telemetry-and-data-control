# main.py
import streamlit as st
import pandas as pd
import psycopg2
from datetime import datetime, timedelta
import requests
import re
import hashlib
import plotly.express as px
import os


# === НАСТРОЙКА ПОДКЛЮЧЕНИЯ К БД ===
def get_db_connection():
    # Если запущено в Docker - используем переменные окружения
    if os.environ.get('DB_HOST'):
        return psycopg2.connect(
            host=os.environ['DB_HOST'],
            port=os.environ['DB_PORT'],
            database=os.environ['DB_NAME'],
            user=os.environ['DB_USER'],
            password=os.environ['DB_PASSWORD']
        )
    # Локальный запуск
    return psycopg2.connect(
        host='localhost',
        port=5432,
        database='data',
        user='postgres',
        password='4845'
    )


# === НАСТРОЙКА СТРАНИЦЫ ===
st.set_page_config(
    page_title="Датчики | Мониторинг",
    page_icon="📡",
    layout="wide",
    initial_sidebar_state="expanded"
)


def hash_password(password):
    return hashlib.sha256(password.encode()).hexdigest()


def check_login(username, password):
    conn = get_db_connection()
    cur = conn.cursor()
    hashed = hash_password(password)
    cur.execute(
        "SELECT id, username, role FROM users WHERE username = %s AND password = %s",
        (username, hashed)
    )
    user = cur.fetchone()
    conn.close()
    if user:
        return {"id": user[0], "username": user[1], "role": user[2]}
    return None


def register_user(username, password, role="user"):
    """Регистрация нового пользователя"""
    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("SELECT id FROM users WHERE username = %s", (username,))
    if cur.fetchone():
        conn.close()
        return False, "Пользователь уже существует"

    hashed = hash_password(password)
    try:
        cur.execute("""
            INSERT INTO users (username, password, role, created_at)
            VALUES (%s, %s, %s, %s)
        """, (username, hashed, role, datetime.now()))
        conn.commit()
        conn.close()
        return True, "Регистрация успешна!"
    except Exception as e:
        conn.close()
        return False, f"Ошибка: {e}"


def init_db_tables():
    """Создаёт все таблицы если их нет"""
    conn = get_db_connection()
    cur = conn.cursor()

    # Таблица users
    cur.execute("""
        CREATE TABLE IF NOT EXISTS users (
            id SERIAL PRIMARY KEY,
            username VARCHAR(100) UNIQUE NOT NULL,
            password VARCHAR(255) NOT NULL,
            role VARCHAR(50) DEFAULT 'user',
            created_at TIMESTAMP DEFAULT NOW()
        )
    """)

    # Таблица sensor_readings
    cur.execute("""
        CREATE TABLE IF NOT EXISTS sensor_readings (
            id SERIAL PRIMARY KEY,
            installation_id INTEGER,
            sensor_id INTEGER,
            sensor_name VARCHAR(100),
            value FLOAT,
            status VARCHAR(50),
            recorded_at TIMESTAMP DEFAULT NOW()
        )
    """)

    # Таблица events
    cur.execute("""
        CREATE TABLE IF NOT EXISTS events (
            id SERIAL PRIMARY KEY,
            event_ts TIMESTAMP DEFAULT NOW(),
            installation_id INTEGER,
            sensor_id INTEGER,
            event_type VARCHAR(50),
            short_text TEXT
        )
    """)

    conn.commit()
    conn.close()
    print("[DEBUG] Таблицы созданы/проверены")


def init_default_admin():
    """Создать админа если нет пользователей"""
    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("SELECT COUNT(*) FROM users")
    count = cur.fetchone()[0]
    conn.close()

    if count == 0:
        register_user("admin", "admin", "admin")
        register_user("user", "user123", "user")


def init_events_table():
    """Создаёт таблицу events если её нет"""
    conn = get_db_connection()
    cur = conn.cursor()

    cur.execute("""
        CREATE TABLE IF NOT EXISTS events (
            id SERIAL PRIMARY KEY,
            event_ts TIMESTAMP,
            installation_id INTEGER,
            sensor_id INTEGER,
            event_type VARCHAR(50),
            short_text TEXT
        )
    """)
    conn.commit()
    conn.close()


def refresh_events_from_sensors():
    """Обновляет события из текущих показаний датчиков, используя пороги из тепловой карты"""
    conn = get_db_connection()
    cur = conn.cursor()

    # Твои пороги из тепловой карты
    thresholds = {
        1: {'warn': 87, 'alarm': 93, 'name': 'Температура', 'unit': '°C'},
        2: {'warn': 4.9, 'alarm': 5.5, 'name': 'Давление', 'unit': 'бар'},
        3: {'warn': 63, 'alarm': 70, 'name': 'Расход', 'unit': 'м³/ч'},
        4: {'warn': 67, 'alarm': 75, 'name': 'Уровень', 'unit': '%'}
    }

    # Очищаем старые события
    cur.execute("DELETE FROM events")
    conn.commit()

    # Получаем последние показания датчиков
    df = get_sensors_with_thresholds()

    if df.empty:
        conn.close()
        return

    created = 0
    for _, row in df.iterrows():
        sensor_id = row['sensor_id']
        value = row['value']
        inst_id = row['installation_id']
        sensor_name = row['sensor_name']
        inst_name = row['installation_name']

        # Получаем пороги для датчика
        thr = thresholds.get(sensor_id)
        if not thr:
            continue

        # ОПРЕДЕЛЯЕМ СТАТУС ПО ПОРОГАМ, а не из API
        if value > thr['alarm']:
            event_type = 'alarm'
            text = f"🚨 АЛАРМ: {sensor_name} = {value:.1f} {thr['unit']} (крит. {thr['alarm']}{thr['unit']}) на {inst_name}"
        elif value > thr['warn']:
            event_type = 'warning'
            text = f"⚠️ ПРЕДУПРЕЖДЕНИЕ: {sensor_name} = {value:.1f} {thr['unit']} (порог {thr['warn']}{thr['unit']}) на {inst_name}"
        else:
            # Норма - не создаём событие
            continue

        cur.execute("""
            INSERT INTO events (event_ts, installation_id, sensor_id, event_type, short_text)
            VALUES (NOW(), %s, %s, %s, %s)
        """, (inst_id, sensor_id, event_type, text))
        created += 1
        print(f"[DEBUG] {event_type}: {sensor_name} = {value} {thr['unit']}")

    conn.commit()
    conn.close()
    print(f"[DEBUG] Всего создано событий: {created}")


def save_sensor_data(installation_id, sensor_id, sensor_name, value, status):
    conn = get_db_connection()
    cur = conn.cursor()

    if isinstance(sensor_id, str):
        numbers = re.findall(r'\d+', sensor_id)
        sensor_id = int(numbers[-1]) if numbers else 1

    cur.execute("""
        INSERT INTO sensor_readings 
        (installation_id, sensor_id, sensor_name, value, status, recorded_at)
        VALUES (%s, %s, %s, %s, %s, %s)
    """, (installation_id, sensor_id, sensor_name, value, status, datetime.now()))
    conn.commit()
    conn.close()


def get_sensors_with_thresholds():
    conn = get_db_connection()
    query = """
        SELECT DISTINCT ON (installation_id, sensor_id) 
            installation_id,
            sensor_id,
            sensor_name,
            value,
            status,
            recorded_at as timestamp
        FROM sensor_readings
        ORDER BY installation_id, sensor_id, recorded_at DESC
    """
    df = pd.read_sql(query, conn)
    conn.close()

    if not df.empty:
        df['installation_name'] = df['installation_id'].apply(lambda x: f"Установка {x}")
        units = {1: '°C', 2: 'бар', 3: 'м³/ч', 4: '%'}
        df['unit'] = df['sensor_id'].map(units)

    return df


def get_events(from_date=None, to_date=None, sort_order="DESC"):
    conn = get_db_connection()
    query = """
            SELECT 
                id,
                event_ts as timestamp,
                installation_id,
                sensor_id,
                event_type as severity,
                short_text as message
            FROM events
            WHERE 1=1
        """
    params = []

    if from_date:
        query += " AND event_ts >= %s"
        params.append(from_date)
    if to_date:
        query += " AND event_ts <= %s"
        params.append(to_date)

    if sort_order == "ASC":
        query += " ORDER BY event_ts ASC"
    elif sort_order == "ALARM_FIRST":
        query += " ORDER BY CASE WHEN event_type = 'alarm' THEN 0 WHEN event_type = 'warning' THEN 1 ELSE 2 END, event_ts DESC"
    else:
        query += " ORDER BY event_ts DESC"

    query += " LIMIT 500"

    df = pd.read_sql(query, conn, params=params if params else None)
    conn.close()

    if not df.empty:
        df['installation_name'] = df['installation_id'].apply(lambda x: f"Установка {x}")
        df['sensor_name'] = df['sensor_id'].apply(lambda x: f"Датчик {x}" if x else "Система")
        df['timestamp'] = pd.to_datetime(df['timestamp']).dt.tz_localize(None)

    return df


def fetch_and_save_from_api():
    try:
        url = "http://5.129.248.80:8001/sensors"
        response = requests.get(url, timeout=10)

        if response.status_code == 200:
            data = response.json()

            for installation in data:
                inst_id = installation['installation_id']

                for sensor in installation.get('sensors', []):
                    save_sensor_data(
                        installation_id=inst_id,
                        sensor_id=sensor['sensor_id'],
                        sensor_name=sensor['sensor_name'],
                        value=sensor['value'],
                        status=sensor['status']
                    )

            # Обновляем события из свежих данных
            refresh_events_from_sensors()

            return True
        return False
    except Exception as e:
        print(f"Ошибка API: {e}")
        return False


def get_all_users():
    """Получить список всех пользователей (только для админа)"""
    conn = get_db_connection()
    df = pd.read_sql("SELECT id, username, role, created_at FROM users ORDER BY id", conn)
    conn.close()
    return df


def delete_user(user_id, current_user_id):
    """Удалить пользователя по ID (нельзя удалить себя)"""
    if user_id == current_user_id:
        return False, "Нельзя удалить самого себя"
    conn = get_db_connection()
    cur = conn.cursor()
    user_id = int(user_id)
    cur.execute("DELETE FROM users WHERE id = %s", (user_id,))
    conn.commit()
    conn.close()
    return True, "Пользователь удалён"


def change_user_role(user_id, new_role, current_user_id, current_user_role):
    """Изменить роль пользователя (нельзя снять админку с себя, нельзя снять админку с другого админа)"""
    if user_id == current_user_id and current_user_role == "admin" and new_role != "admin":
        return False, "Нельзя снять админку с самого себя"

    conn = get_db_connection()
    cur = conn.cursor()
    cur.execute("SELECT role FROM users WHERE id = %s", (int(user_id),))
    target_role = cur.fetchone()
    conn.close()

    if target_role and target_role[0] == "admin" and new_role != "admin":
        return False, "Нельзя снять админку с другого администратора"

    conn = get_db_connection()
    cur = conn.cursor()
    user_id = int(user_id)
    cur.execute("UPDATE users SET role = %s WHERE id = %s", (new_role, user_id))
    conn.commit()
    conn.close()
    return True, f"Роль пользователя изменена на {new_role}"


def get_system_stats():
    """Получить статистику системы (только для админа)"""
    conn = get_db_connection()
    stats = {}

    cur = conn.cursor()
    cur.execute("SELECT COUNT(*) FROM users")
    stats['total_users'] = cur.fetchone()[0]

    cur.execute("SELECT COUNT(*) FROM sensor_readings")
    stats['total_readings'] = cur.fetchone()[0]

    cur.execute("SELECT COUNT(*) FROM events")
    stats['total_events'] = cur.fetchone()[0]

    cur.execute("SELECT COUNT(DISTINCT installation_id) FROM sensor_readings")
    stats['total_installations'] = cur.fetchone()[0]

    conn.close()
    return stats


def get_daily_events():
    """Получить количество событий по дням для графика"""
    conn = get_db_connection()
    df = pd.read_sql("""
        SELECT DATE(event_ts) as date, COUNT(*) as count
        FROM events
        GROUP BY DATE(event_ts)
        ORDER BY date DESC
        LIMIT 30
    """, conn)
    conn.close()
    return df


# === CSS ===
st.markdown("""
<style>
    .main {
        background: linear-gradient(135deg, #f5f7fa 0%, #c3cfe2 100%);
    }
    .dashboard-title {
        font-size: 28px;
        font-weight: bold;
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
        margin-bottom: 20px;
    }
    .metric-card {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        border-radius: 12px;
        padding: 15px;
        color: white;
        text-align: center;
    }
    .metric-value { font-size: 32px; font-weight: bold; }
    .metric-label { font-size: 14px; opacity: 0.9; }

    .status-normal { display: inline-block; width: 12px; height: 12px; border-radius: 50%; background-color: #4CAF50; animation: pulse-green 2s infinite; }
    .status-warning { display: inline-block; width: 12px; height: 12px; border-radius: 50%; background-color: #FFC107; animation: pulse-yellow 2s infinite; }
    .status-alarm { display: inline-block; width: 12px; height: 12px; border-radius: 50%; background-color: #F44336; animation: pulse-red 1s infinite; }

    @keyframes pulse-green {
        0% { box-shadow: 0 0 0 0 rgba(76, 175, 80, 0.7); }
        70% { box-shadow: 0 0 0 10px rgba(76, 175, 80, 0); }
        100% { box-shadow: 0 0 0 0 rgba(76, 175, 80, 0); }
    }
    @keyframes pulse-yellow {
        0% { box-shadow: 0 0 0 0 rgba(255, 193, 7, 0.7); }
        70% { box-shadow: 0 0 0 10px rgba(255, 193, 7, 0); }
        100% { box-shadow: 0 0 0 0 rgba(255, 193, 7, 0); }
    }
    @keyframes pulse-red {
        0% { box-shadow: 0 0 0 0 rgba(244, 67, 54, 0.7); }
        70% { box-shadow: 0 0 0 10px rgba(244, 67, 54, 0); }
        100% { box-shadow: 0 0 0 0 rgba(244, 67, 54, 0); }
    }

    .sensor-card {
        background: white;
        border-radius: 12px;
        padding: 15px;
        margin: 10px 0;
        border-left: 4px solid;
        transition: all 0.3s ease;
    }
    .sensor-card:hover { transform: translateX(5px); box-shadow: 0 4px 12px rgba(0,0,0,0.1); }
    .sensor-value { font-size: 24px; font-weight: bold; }
    .sensor-unit { font-size: 14px; color: #666; }

    .heatmap-table { width: 100%; border-collapse: collapse; }
    .heatmap-table th { background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: white; padding: 12px; text-align: center; font-weight: bold; }
    .heatmap-table td { padding: 12px; text-align: center; border: 1px solid #e0e0e0; transition: all 0.3s ease; }
    .heatmap-table td:hover { transform: scale(1.02); font-weight: bold; }

    .stButton > button { background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: white; border: none; border-radius: 8px; padding: 10px 24px; font-weight: bold; transition: all 0.3s ease; }
    .stButton > button:hover { transform: translateY(-2px); box-shadow: 0 4px 12px rgba(102, 126, 234, 0.4); }

    .css-1d391kg { background: linear-gradient(180deg, #1a1a2e 0%, #16213e 100%); }
    .custom-progress { height: 6px; border-radius: 3px; background: linear-gradient(90deg, #4CAF50, #FFC107, #F44336); margin-top: 8px; }
    .event-separator { height: 1px; background: #e0e0e0; margin: 10px 0; }
</style>
""", unsafe_allow_html=True)


# === АВТОРИЗАЦИЯ ===
def login_page():
    _, col_main, _ = st.columns([1, 2, 1])

    with col_main:
        st.markdown(
            '<div style="text-align: center; font-size: 24px; font-weight: 600; margin-bottom: 2rem;">👋 Добро пожаловать</div>',
            unsafe_allow_html=True)

        tab1, tab2 = st.tabs(["🔐 Вход", "📝 Регистрация"])

        with tab1:
            with st.form("login_form", clear_on_submit=False):
                username = st.text_input("Логин", placeholder="Введите ваш логин", key="login_username")
                password = st.text_input("Пароль", type="password", placeholder="••••••••", key="login_password")

                submitted = st.form_submit_button("Войти", use_container_width=True)

                if submitted:
                    if username and password:
                        user = check_login(username, password)
                        if user:
                            st.session_state['authenticated'] = True
                            st.session_state['user'] = user
                            st.rerun()
                        else:
                            st.error("❌ Неверный логин или пароль")
                    else:
                        st.warning("Заполните все поля")

        with tab2:
            with st.form("register_form", clear_on_submit=False):
                new_username = st.text_input("Логин", placeholder="Придумайте логин", key="reg_username")
                new_password = st.text_input("Пароль", type="password", placeholder="Минимум 4 символа",
                                             key="reg_password")
                confirm_password = st.text_input("Подтвердите пароль", type="password", placeholder="Повторите пароль",
                                                 key="reg_confirm")

                submitted_reg = st.form_submit_button("Зарегистрироваться", use_container_width=True)

                if submitted_reg:
                    if not new_username or not new_password:
                        st.error("Заполните все поля")
                    elif new_password != confirm_password:
                        st.error("Пароли не совпадают")
                    elif len(new_password) < 4:
                        st.error("Пароль должен быть не менее 4 символов")
                    else:
                        success, message = register_user(new_username, new_password, "user")
                        if success:
                            st.success(message)
                            user = check_login(new_username, new_password)
                            if user:
                                st.session_state['authenticated'] = True
                                st.session_state['user'] = user
                                st.rerun()
                        else:
                            st.error(message)


def admin_panel():
    """Админ-панель (только для пользователей с ролью admin)"""
    st.markdown("### 👑 Админ-панель")

    tabs = st.tabs(["📊 Статистика", "👥 Пользователи", "🗑 Очистка"])

    # Вкладка статистики
    with tabs[0]:
        st.markdown("#### 📈 Системная статистика")

        stats = get_system_stats()

        col1, col2, col3, col4 = st.columns(4)
        with col1:
            st.metric("👥 Пользователей", stats['total_users'])
        with col2:
            st.metric("📡 Показаний датчиков", stats['total_readings'])
        with col3:
            st.metric("📋 Событий", stats['total_events'])
        with col4:
            st.metric("🏭 Установок", stats['total_installations'])

        # График активности по дням
        st.markdown("#### 📅 Активность (события по дням)")
        events_df = get_daily_events()
        if not events_df.empty:
            events_df = events_df.sort_values('date')
            if len(events_df) > 1:
                fig = px.line(events_df, x='date', y='count', title='Количество событий по дням', markers=True)
                fig.update_layout(height=400)
                st.plotly_chart(fig, use_container_width=True)
            else:
                st.info(f"ℹ️ Найдено {len(events_df)} событий за сегодня")
        else:
            st.info("ℹ️ Нет данных о событиях. События создаются автоматически из данных датчиков.")

    # Вкладка управления пользователями
    with tabs[1]:
        st.markdown("#### 👥 Управление пользователями")

        users_df = get_all_users()
        current_user_id = st.session_state['user']['id']
        current_user_role = st.session_state['user']['role']

        if not users_df.empty:
            st.dataframe(users_df, use_container_width=True)

            st.markdown("---")
            st.markdown("#### ✏️ Действия с пользователями")

            col1, col2 = st.columns(2)

            with col1:
                st.markdown("**🗑 Удаление пользователя**")
                user_to_delete = st.selectbox(
                    "Выберите пользователя для удаления",
                    users_df['username'].tolist(),
                    key="delete_user"
                )
                if st.button("🗑 Удалить пользователя", key="delete_btn"):
                    user_id = int(users_df[users_df['username'] == user_to_delete]['id'].values[0])
                    if user_id == current_user_id:
                        st.error("❌ Нельзя удалить самого себя")
                    else:
                        success, message = delete_user(user_id, current_user_id)
                        if success:
                            st.success(message)
                            st.rerun()
                        else:
                            st.error(message)

            with col2:
                st.markdown("**🔄 Смена роли**")
                user_to_change = st.selectbox(
                    "Выберите пользователя",
                    users_df['username'].tolist(),
                    key="change_role_user"
                )
                new_role = st.selectbox(
                    "Новая роль",
                    ["user", "admin"],
                    key="new_role"
                )
                if st.button("🔄 Сменить роль", key="change_role_btn"):
                    user_id = int(users_df[users_df['username'] == user_to_change]['id'].values[0])
                    target_user_role = users_df[users_df['username'] == user_to_change]['role'].values[0]

                    if user_id == current_user_id and current_user_role == "admin" and new_role != "admin":
                        st.error("❌ Нельзя снять админку с самого себя")
                    elif target_user_role == "admin" and new_role != "admin":
                        st.error("❌ Нельзя снять админку с другого администратора")
                    elif target_user_role == new_role:
                        st.warning(f"У пользователя {user_to_change} уже роль {new_role}")
                    else:
                        success, message = change_user_role(user_id, new_role, current_user_id, current_user_role)
                        if success:
                            st.success(message)
                            st.rerun()
                        else:
                            st.error(message)

    # Вкладка очистки
    with tabs[2]:
        st.markdown("#### 🗑 Очистка данных")
        st.warning("⚠️ Внимание! Эти действия необратимы.")

        col1, col2 = st.columns(2)

        with col1:
            if st.button("🧹 Очистить журнал событий", key="clear_events_btn"):
                conn = get_db_connection()
                cur = conn.cursor()
                cur.execute("DELETE FROM events")
                conn.commit()
                conn.close()
                st.success("Журнал событий очищен")
                st.rerun()

        with col2:
            if st.button("📡 Очистить данные датчиков", key="clear_sensors_btn"):
                conn = get_db_connection()
                cur = conn.cursor()
                cur.execute("DELETE FROM sensor_readings")
                conn.commit()
                conn.close()
                st.success("Данные датчиков очищены")
                st.rerun()


def logout():
    if st.button("🚪 Выйти", use_container_width=True):
        st.session_state['authenticated'] = False
        st.session_state['user'] = None
        st.rerun()


# === ИНИЦИАЛИЗАЦИЯ ===
if 'authenticated' not in st.session_state:
    st.session_state['authenticated'] = False
    st.session_state['user'] = None

# Создаём таблицы перед всем остальным
init_db_tables()
init_default_admin()
init_events_table()
refresh_events_from_sensors()

if not st.session_state['authenticated']:
    login_page()
    st.stop()

# === ОСНОВНОЕ ПРИЛОЖЕНИЕ ===
fetch_and_save_from_api()

st.markdown('<div class="dashboard-title">📡 Телеметрия и контроль продуктовых данных</div>', unsafe_allow_html=True)

# Сайдбар
with st.sidebar:
    if st.session_state['user']['role'] == 'admin':
        st.markdown("👑 **Режим администратора**")
    else:
        st.markdown("👤 **Режим пользователя**")

    st.markdown(f"**Имя:** {st.session_state['user']['username']}")
    st.markdown(f"**Роль:** {st.session_state['user']['role']}")

    logout()
    st.markdown("---")

    st.markdown("### 🎮 Управление")
    st.markdown("---")

    st.markdown("**🔗 URL API телеметрии**")
    st.code("http://5.129.248.80:8001", language="url")

    st.markdown("---")

    if st.button("📡 Собрать данные из API", use_container_width=True):
        with st.spinner("Сбор данных..."):
            if fetch_and_save_from_api():
                st.success("Данные собраны!")
                st.rerun()
            else:
                st.error("Ошибка сбора")

    col1, col2 = st.columns(2)
    with col1:
        if st.button("🔄 Обновить", use_container_width=True):
            st.cache_data.clear()
            st.rerun()
    with col2:
        auto_update = st.checkbox("Автообновление", value=False)

    st.markdown("---")

    all_events = get_events()
    st.markdown(f"""
    <div class="metric-card">
        <div class="metric-value">{len(all_events)}</div>
        <div class="metric-label">Событий в журнале</div>
    </div>
    """, unsafe_allow_html=True)

    st.markdown('<div style="height: 15px;"></div>', unsafe_allow_html=True)

    if st.session_state['user']['role'] == 'admin':
        if st.button("🗑 Очистить журнал", use_container_width=True):
            conn = get_db_connection()
            cur = conn.cursor()
            cur.execute("DELETE FROM events")
            conn.commit()
            conn.close()
            refresh_events_from_sensors()
            st.success("Журнал очищен и обновлён")
            st.rerun()

    st.markdown("---")
    st.markdown("### 📊 Фильтры")

    sensors_df = get_sensors_with_thresholds()
    if not sensors_df.empty:
        installations = sensors_df['installation_name'].unique()
        selected_installs = st.multiselect("Установки", installations, default=list(installations))

        all_sensors = sensors_df['sensor_name'].unique()
        selected_sensors = st.multiselect("Датчики", all_sensors, default=list(all_sensors))

# Админ-панель (только для админа)
if st.session_state['user']['role'] == 'admin':
    admin_panel()
    st.markdown("---")

# Получаем данные
sensors_df = get_sensors_with_thresholds()

if not sensors_df.empty:
    filtered_df = sensors_df.copy()

    if 'selected_installs' in locals() and selected_installs:
        filtered_df = filtered_df[filtered_df['installation_name'].isin(selected_installs)]

    if 'selected_sensors' in locals() and selected_sensors:
        filtered_df = filtered_df[filtered_df['sensor_name'].isin(selected_sensors)]

    # === ТЕПЛОВАЯ КАРТА ===
    st.markdown("### 🔥 Тепловая карта состояния датчиков")

    heatmap_data = sensors_df.pivot_table(
        index='installation_name',
        columns='sensor_name',
        values='value',
        aggfunc='first'
    )


    def get_cell_color(value, sensor_id):
        if pd.isna(value):
            return '#f0f0f0'
        thresholds = {1: {'warn': 87, 'alarm': 93}, 2: {'warn': 4.9, 'alarm': 5.5},
                      3: {'warn': 63, 'alarm': 70}, 4: {'warn': 67, 'alarm': 75}}
        thr = thresholds.get(sensor_id, {'warn': 100, 'alarm': 120})
        if value > thr['alarm'] or value < (thr['warn'] * 0.5):
            return '#FF6B6B'
        elif value > thr['warn']:
            return '#FFD700'
        return '#90EE90'


    html_table = '<table class="heatmap-table">'
    html_table += '<thead><tr><th>Установка / Датчик</th>'

    for col in heatmap_data.columns:
        html_table += f'<th>{col}</th>'
    html_table += '</tr></thead><tbody>'

    for idx in heatmap_data.index:
        html_table += f'<tr><td style="font-weight: bold; background: #f5f5f5;">{idx}</td>'
        for col in heatmap_data.columns:
            value = heatmap_data.loc[idx, col]
            if pd.isna(value):
                html_table += '<td>—</td>'
            else:
                sensor_id = {'Датчик 1': 1, 'Датчик 2': 2, 'Датчик 3': 3, 'Датчик 4': 4}.get(col, 1)
                color = get_cell_color(value, sensor_id)
                unit = {1: '°C', 2: 'бар', 3: 'м³/ч', 4: '%'}.get(sensor_id, '')
                html_table += f'<td style="background-color: {color}; font-weight: bold;">{value:.1f} {unit}</td>'
        html_table += '</tr>'
    html_table += '</tbody></table>'
    st.markdown(html_table, unsafe_allow_html=True)

    col1, col2, col3 = st.columns(3)
    with col1:
        st.markdown('<div style="text-align: center;"><span class="status-normal"></span> 🟢 Норма</div>',
                    unsafe_allow_html=True)
    with col2:
        st.markdown('<div style="text-align: center;"><span class="status-warning"></span> 🟡 Предупреждение</div>',
                    unsafe_allow_html=True)
    with col3:
        st.markdown('<div style="text-align: center;"><span class="status-alarm"></span> 🔴 Аларм</div>',
                    unsafe_allow_html=True)

    st.markdown("---")

    # === ДЕРЕВО ДАТЧИКОВ ===
    st.markdown("### 🌳 Структура датчиков по установкам")

    for installation in filtered_df['installation_name'].unique():
        with st.expander(f"🏭 {installation}", expanded=True):
            inst_data = filtered_df[filtered_df['installation_name'] == installation]
            for _, row in inst_data.iterrows():
                status_color = {'normal': '#4CAF50', 'warning': '#FFC107', 'alarm': '#F44336'}.get(row['status'],
                                                                                                   '#666')
                status_class = {'normal': 'status-normal', 'warning': 'status-warning', 'alarm': 'status-alarm'}.get(
                    row['status'], 'status-normal')
                unit = row.get('unit', '')

                st.markdown(f"""
                <div class="sensor-card" style="border-left-color: {status_color};">
                    <div style="display: flex; justify-content: space-between; align-items: center;">
                        <div>
                            <span class="{status_class}"></span>
                            <strong style="margin-left: 8px;">{row['sensor_name']}</strong>
                        </div>
                        <div style="text-align: right;">
                            <span class="sensor-value">{row['value']:.2f}</span>
                            <span class="sensor-unit">{unit}</span>
                        </div>
                    </div>
                    <div class="custom-progress" style="width: {min(100, (row['value'] / 120) * 100)}%;"></div>
                </div>
                """, unsafe_allow_html=True)

    st.markdown("---")

    # === ЛЕНТА СОБЫТИЙ ===
    st.markdown("### 📋 Лента событий и алармов")

    col1, col2, col3 = st.columns(3)
    with col1:
        date_from = st.date_input("📅 С даты", datetime.now() - timedelta(days=7))
    with col2:
        date_to = st.date_input("📅 По дату", datetime.now())
    with col3:
        sort_option = st.selectbox("📊 Сортировка", ["Новые сначала", "Старые сначала", "Алармы сначала"])

    time_from = datetime.combine(date_from, datetime.min.time())
    time_to = datetime.combine(date_to, datetime.max.time())

    sort_map = {"Новые сначала": "DESC", "Старые сначала": "ASC", "Алармы сначала": "ALARM_FIRST"}

    events = get_events(time_from, time_to, sort_map.get(sort_option, "DESC"))

    if not events.empty:
        st.caption(f"📌 Всего событий: {len(events)}")

        for i, (_, event) in enumerate(events.iterrows()):
            if event['severity'] == 'alarm':
                icon = "🚨"
                color = "#F44336"
                bg = "#FFEBEE"
            elif event['severity'] == 'warning':
                icon = "⚠️"
                color = "#FFC107"
                bg = "#FFF8E1"
            else:
                icon = "ℹ️"
                color = "#2196F3"
                bg = "#E3F2FD"

            st.markdown(f"""
            <div style="background: {bg}; border-radius: 8px; padding: 12px; margin: 8px 0; border-left: 4px solid {color}; box-shadow: 0 2px 4px rgba(0,0,0,0.1);">
                <div style="display: flex; justify-content: space-between; align-items: center;">
                    <div>
                        <span style="font-size: 20px;">{icon}</span>
                        <strong>{event['installation_name']}</strong>
                        <span style="color: #666; margin-left: 8px;">{event['timestamp'].strftime('%Y-%m-%d %H:%M:%S')}</span>
                    </div>
                    <span style="background: {color}; color: white; padding: 4px 12px; border-radius: 20px; font-size: 12px;">
                        {event['severity'].upper()}
                    </span>
                </div>
                <div style="margin-top: 8px; color: #333;">
                    {event['message']}
                </div>
            </div>
            """, unsafe_allow_html=True)

            if i < len(events) - 1:
                st.markdown('<div class="event-separator"></div>', unsafe_allow_html=True)
    else:
        st.info("📭 Нет событий за выбранный период")

    # === ЭКСПОРТ ===
    st.markdown("### 📥 Выгрузка отчёта")

    col1, col2 = st.columns(2)
    with col1:
        exp_from = st.date_input("С даты (экспорт)", datetime.now() - timedelta(days=30))
    with col2:
        exp_to = st.date_input("По дату (экспорт)", datetime.now())

    if st.button("📊 Экспорт отчёта в Excel", use_container_width=True):
        exp_events = get_events(
            datetime.combine(exp_from, datetime.min.time()),
            datetime.combine(exp_to, datetime.max.time()),
            "DESC"
        )

        if not exp_events.empty:
            from io import BytesIO
            from openpyxl.styles import PatternFill, Alignment, Font
            from openpyxl.utils import get_column_letter

            output = BytesIO()

            # Подготовка данных
            export_df = exp_events[['timestamp', 'installation_name', 'severity', 'message']].copy()
            export_df.columns = ['Дата', 'Установка', 'Тип', 'Событие']

            # Маппинг типов
            severity_map = {'alarm': '🚨 АЛАРМ', 'warning': '⚠️ ПРЕДУПРЕЖДЕНИЕ', 'info': 'ℹ️ ИНФО'}
            export_df['Тип'] = export_df['Тип'].map(severity_map)

            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                export_df.to_excel(writer, index=False, sheet_name='События')

                ws = writer.sheets['События']

                # Настройка ширины колонок
                col_widths = {'A': 20, 'B': 18, 'C': 20, 'D': 50}
                for col, width in col_widths.items():
                    ws.column_dimensions[col].width = width

                # Стиль для заголовков
                header_font = Font(bold=True, size=11)
                header_align = Alignment(horizontal='center', vertical='center')

                for col_num in range(1, 5):
                    cell = ws.cell(row=1, column=col_num)
                    cell.font = header_font
                    cell.alignment = header_align

                # Цвета для колонки Тип (3-я колонка)
                alarm_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
                warning_fill = PatternFill(start_color="FFCC66", end_color="FFCC66", fill_type="solid")
                info_fill = PatternFill(start_color="99CC99", end_color="99CC99", fill_type="solid")

                # Выравнивание
                align_center = Alignment(horizontal='center', vertical='center')
                align_left = Alignment(horizontal='left', vertical='center', wrap_text=True)

                for row_num in range(2, len(export_df) + 2):
                    severity = ws.cell(row=row_num, column=3).value

                    if 'АЛАРМ' in str(severity):
                        fill = alarm_fill
                    elif 'ПРЕДУПРЕЖДЕНИЕ' in str(severity):
                        fill = warning_fill
                    else:
                        fill = info_fill

                    # Закрашиваем только 3-ю колонку (Тип)
                    ws.cell(row=row_num, column=3).fill = fill

                    # Выравнивание
                    ws.cell(row=row_num, column=1).alignment = align_center
                    ws.cell(row=row_num, column=2).alignment = align_center
                    ws.cell(row=row_num, column=3).alignment = align_center
                    ws.cell(row=row_num, column=4).alignment = align_left

            st.download_button(
                label="📥 Скачать Excel",
                data=output.getvalue(),
                file_name=f"events_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            st.success("✅ Отчёт готов!")
        else:
            st.warning("Нет событий за выбранный период")
else:
    st.warning("⚠️ Нет данных от датчиков. Нажмите 'Собрать данные из API' в боковой панели.")

if auto_update:
    import time

    time.sleep(30)
    st.rerun()